from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_wtf import CSRFProtect
from werkzeug.security import generate_password_hash, check_password_hash
import json
import os
from datetime import datetime
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import calendar
import locale

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-here')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///work_hours.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
csrf = CSRFProtect(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Constants
HOURLY_RATE = 14
OVERTIME_MULTIPLIER = 1.5

# Models
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    entries = db.relationship('WorkEntry', backref='user', lazy=True)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class WorkEntry(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False)
    day = db.Column(db.String(20), nullable=False)
    start_time = db.Column(db.Time, nullable=False)
    end_time = db.Column(db.Time, nullable=False)
    total_hours = db.Column(db.Float, nullable=False)
    overtime_hours = db.Column(db.Float, nullable=False)
    pay = db.Column(db.Float, nullable=False)
    note = db.Column(db.String(200))
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Routes
@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        remember = request.form.get('remember') == 'on'
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            login_user(user, remember=remember)
            return redirect(url_for('index'))
        flash('Invalid username or password')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/api/entries', methods=['GET'])
@login_required
def get_entries():
    entries = WorkEntry.query.filter_by(user_id=current_user.id).all()
    return jsonify([{
        'date': entry.date.strftime('%Y-%m-%d'),
        'day': entry.day,
        'start': entry.start_time.strftime('%H:%M'),
        'end': entry.end_time.strftime('%H:%M'),
        'totalHours': entry.total_hours,
        'overtimeHours': entry.overtime_hours,
        'pay': entry.pay,
        'note': entry.note
    } for entry in entries])

@app.route('/api/entries', methods=['POST'])
@login_required
def add_entry():
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'لم يتم استلام البيانات'}), 400

        # Validate required fields
        required_fields = ['date', 'day', 'start', 'end', 'totalHours', 'overtimeHours', 'pay']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'حقل {field} مطلوب'}), 400

        # Parse and validate date
        try:
            date = datetime.strptime(data['date'], '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'صيغة التاريخ غير صحيحة'}), 400

        # Check if entry exists for this date
        existing_entry = WorkEntry.query.filter_by(
            user_id=current_user.id,
            date=date
        ).first()
        
        if existing_entry:
            return jsonify({'error': 'يوجد بالفعل تسجيل لهذا التاريخ'}), 400

        # Parse and validate times
        try:
            start_time = datetime.strptime(data['start'], '%H:%M').time()
            end_time = datetime.strptime(data['end'], '%H:%M').time()
        except ValueError:
            return jsonify({'error': 'صيغة الوقت غير صحيحة'}), 400

        # Validate time order
        if end_time <= start_time:
            return jsonify({'error': 'وقت الانتهاء يجب أن يكون بعد وقت البدء'}), 400

        # Create new entry
        entry = WorkEntry(
            date=date,
            day=data['day'],
            start_time=start_time,
            end_time=end_time,
            total_hours=float(data['totalHours']),
            overtime_hours=float(data['overtimeHours']),
            pay=float(data['pay']),
            note=data.get('note', ''),
            user_id=current_user.id
        )
        
        db.session.add(entry)
        db.session.commit()
        
        return jsonify({
            'message': 'تمت الإضافة بنجاح',
            'entry': {
                'date': entry.date.strftime('%Y-%m-%d'),
                'day': entry.day,
                'start': entry.start_time.strftime('%H:%M'),
                'end': entry.end_time.strftime('%H:%M'),
                'totalHours': entry.total_hours,
                'overtimeHours': entry.overtime_hours,
                'pay': entry.pay,
                'note': entry.note
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'حدث خطأ أثناء الإضافة'}), 500

@app.route('/api/entries/<date>', methods=['DELETE'])
@login_required
def delete_entry(date):
    try:
        entry = WorkEntry.query.filter_by(user_id=current_user.id, date=date).first()
        if not entry:
            return jsonify({'error': 'لم يتم العثور على السجل'}), 404
        
        db.session.delete(entry)
        db.session.commit()
        return jsonify({'message': 'تم حذف السجل بنجاح'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'حدث خطأ أثناء حذف السجل'}), 500

@app.route('/api/entries/<date>', methods=['PUT'])
@login_required
def edit_entry(date):
    try:
        entry = WorkEntry.query.filter_by(user_id=current_user.id, date=date).first()
        if not entry:
            return jsonify({'error': 'لم يتم العثور على السجل'}), 404
        
        data = request.get_json()
        
        # التحقق من البيانات المطلوبة
        required_fields = ['start', 'end', 'totalHours', 'overtimeHours', 'pay', 'note']
        if not all(field in data for field in required_fields):
            return jsonify({'error': 'جميع الحقول مطلوبة'}), 400
        
        # تحديث البيانات
        entry.start_time = datetime.strptime(data['start'], '%H:%M').time()
        entry.end_time = datetime.strptime(data['end'], '%H:%M').time()
        entry.total_hours = float(data['totalHours'])
        entry.overtime_hours = float(data['overtimeHours'])
        entry.pay = float(data['pay'])
        entry.note = data['note']
        
        db.session.commit()
        return jsonify({'message': 'تم تحديث السجل بنجاح'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'حدث خطأ أثناء تحديث السجل'}), 500

@app.route('/api/entries/search/<date>', methods=['GET'])
@login_required
def search_entries(date):
    entries = WorkEntry.query.filter_by(
        user_id=current_user.id,
        date=datetime.strptime(date, '%Y-%m-%d').date()
    ).all()
    
    return jsonify([{
        'date': entry.date.strftime('%Y-%m-%d'),
        'day': entry.day,
        'start': entry.start_time.strftime('%H:%M'),
        'end': entry.end_time.strftime('%H:%M'),
        'totalHours': entry.total_hours,
        'overtimeHours': entry.overtime_hours,
        'pay': entry.pay,
        'note': entry.note
    } for entry in entries])

@app.route('/export')
def export_excel():
    try:
        # إنشاء ملف إكسل جديد
        wb = Workbook()
        ws = wb.active
        ws.title = "ساعات العمل"
        
        # تعريف الأنماط
        header_font = Font(name='Arial', size=12, bold=True)
        header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # عناوين الأعمدة
        headers = [
            'التاريخ',
            'اليوم',
            'وقت البدء',
            'وقت الانتهاء',
            'عدد الساعات',
            'الساعات الإضافية',
            'قيمة الساعة الإضافية',
            'الدفع الكلي',
            'مجموع الراتب',
            'ملاحظات'
        ]
        
        # كتابة العناوين
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            # تعيين عرض العمود
            ws.column_dimensions[get_column_letter(col)].width = 15
            
        # قراءة البيانات من قاعدة البيانات
        entries = WorkEntry.query.filter_by(user_id=current_user.id).all()
        
        # كتابة البيانات
        for row, entry in enumerate(entries, 2):
            # تحويل التاريخ إلى اليوم بالعربي
            date_obj = datetime.strptime(entry.date.strftime('%Y-%m-%d'), '%Y-%m-%d')
            day_name = calendar.day_name[date_obj.weekday()]
            
            row_data = [
                entry.date.strftime('%Y-%m-%d'),
                day_name,
                entry.start_time.strftime('%H:%M'),
                entry.end_time.strftime('%H:%M'),
                f"{entry.total_hours:.2f}",
                f"{entry.overtime_hours:.2f}",
                f"{(HOURLY_RATE * OVERTIME_MULTIPLIER):.2f}",
                f"{entry.pay:.2f}",
                f"{entry.pay:.2f}",
                entry.note
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                
                # تنسيق خاص للأرقام
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
        
        # حفظ الملف مؤقتاً
        from io import BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'ساعات_العمل_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
        )
        
    except Exception as e:
        print(f"Error exporting Excel: {str(e)}")
        return jsonify({'error': 'حدث خطأ أثناء تصدير الملف'}), 500

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if User.query.filter_by(username=username).first():
            flash('اسم المستخدم موجود مسبقاً')
            return redirect(url_for('register'))
        
        user = User(username=username)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        
        flash('تم إنشاء الحساب بنجاح')
        return redirect(url_for('login'))
    return render_template('register.html')

# Create tables
with app.app_context():
    db.create_all()
    
    # Create default admin user if not exists
    if not User.query.filter_by(username='admin').first():
        admin = User(username='admin')
        admin.set_password('admin123')
        db.session.add(admin)
        db.session.commit()

# تعيين اللغة العربية
try:
    locale.setlocale(locale.LC_ALL, 'ar_SA.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_ALL, 'ar_AE.UTF-8')
    except:
        pass

if __name__ == '__main__':
    app.run(debug=True) 